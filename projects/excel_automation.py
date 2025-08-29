import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.chart import BarChart, Reference

def automate_excel(input_file, output_file):
    # Read Excel file into a DataFrame
    df = pd.read_excel(input_file, sheet_name='Sheet1')
    
    # Perform calculations: Add a new column for Total Sales
    df['Total Sales'] = df['Units Sold'] * df['Price Per Unit']
    
    # Write DataFrame to a new Excel file
    df.to_excel(output_file, sheet_name='Report', index=False)
    
    # Load workbook for formatting and chart creation
    wb = load_workbook(output_file)
    ws = wb['Report']
    
    # Format header row: Bold and centered
    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # Format Total Sales column as currency
    for row in ws.iter_rows(min_row=2, min_col=ws.max_column, max_col=ws.max_column):
        for cell in row:
            cell.number_format = '$#,##0.00'
    
    # Create a bar chart for Total Sales
    chart = BarChart()
    chart.title = "Total Sales by Product"
    chart.x_axis.title = "Product"
    chart.y_axis.title = "Total Sales ($)"
    data = Reference(ws, min_col=ws.max_column, min_row=2, max_row=ws.max_row)
    categories = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    ws.add_chart(chart, "E2")
    
    # Save the formatted workbook
    wb.save(output_file)
    print(f"Excel automation complete! Output saved to {output_file}")

# Example usage
if __name__ == "__main__":
    input_file = "sales_data.xlsx"
    output_file = "sales_report.xlsx"
    automate_excel(input_file, output_file)